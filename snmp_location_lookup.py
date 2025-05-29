#!/usr/bin/env python3
"""
SNMP Location Lookup Script

This script reads device names from an Excel file, queries the LibreNMS API for each device,
updates the Excel file with the API response data, and checks if the location format is compliant.
"""

import os
import sys
import json
import re
import argparse
import requests
import pandas as pd
import socket
import subprocess
from urllib3.exceptions import InsecureRequestWarning

# Suppress only the specific InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

class LibreNMSClient:
    """Client for interacting with the LibreNMS API"""
    
    def __init__(self, api_url, api_token, verify_ssl=False):
        """
        Initialize the LibreNMS API client
        
        Args:
            api_url (str): Base URL for the LibreNMS API
            api_token (str): API token for authentication
            verify_ssl (bool): Whether to verify SSL certificates
        """
        self.api_url = api_url.rstrip('/')
        self.headers = {'X-Auth-Token': api_token}
        self.verify_ssl = verify_ssl
    
    def get_device_info(self, hostname):
        """
        Get device information from the LibreNMS API
        
        Args:
            hostname (str): Hostname of the device to query
            
        Returns:
            dict: Device information or None if not found
        """
        url = f"{self.api_url}/api/v0/devices/{hostname}"
        try:
            response = requests.get(url, headers=self.headers, verify=self.verify_ssl, timeout=30)
            if response.status_code == 200:
                data = response.json()
                if data.get('status') == 'ok' and data.get('devices') and len(data['devices']) > 0:
                    return data['devices'][0]
                else:
                    print(f"  API returned no device data for {hostname}")
            else:
                print(f"  API returned status code {response.status_code} for {hostname}")
                if response.status_code == 401:
                    print("  Authentication error. Check your API token.")
                elif response.status_code == 404:
                    print("  Device not found in LibreNMS.")
            return None
        except requests.exceptions.Timeout:
            print(f"Error: API request timed out for {hostname}")
            return None
        except requests.exceptions.ConnectionError:
            print(f"Error: Could not connect to API server at {self.api_url}")
            return None
        except Exception as e:
            print(f"Error querying API for {hostname}: {str(e)}")
            return None

def build_expected_location(row, df_columns, format_template):
    """
    Build the expected location string from row data and format template
    
    Args:
        row (pandas.Series): Row data from Excel
        df_columns (list): List of column names from the DataFrame
        format_template (str): Format template with column references (e.g., $B.$C$E.$F)
        
    Returns:
        str: Expected location string
    """
    if not format_template:
        return None
        
    # Replace column references with actual values
    result = format_template
    
    # Find all column references (e.g., $B, $C, etc.)
    col_refs = re.findall(r'\$([A-Z]+)', format_template)
    
    for col_ref in col_refs:
        # Convert column letter to column index (0-based)
        col_idx = ord(col_ref) - ord('A')
        
        # Check if the column index is valid
        if col_idx < 0 or col_idx >= len(df_columns):
            print(f"Warning: Column {col_ref} (index {col_idx}) is out of range. Available columns: {len(df_columns)}")
            # Replace with empty string to avoid template syntax in output
            result = result.replace(f'${col_ref}', '')
            continue
        
        # Get the column name and value
        col_name = df_columns[col_idx]
        
        # Handle NaN values and convert to string
        if pd.isna(row[col_name]):
            col_value = ''
        else:
            # Convert to string and strip whitespace
            col_value = str(row[col_name]).strip()
        
        # Replace column reference with its value
        result = result.replace(f'${col_ref}', col_value)
    
    # Clean up any potential double periods from empty values
    result = re.sub(r'\.\s*\.', '.', result)
    result = re.sub(r'^\.|\.$', '', result)  # Remove leading/trailing periods
    
    return result

def perform_dns_lookup(hostname):
    """
    Perform DNS lookup for a hostname
    
    Args:
        hostname (str): Hostname to lookup
        
    Returns:
        tuple: (IP address, status message)
    """
    try:
        # Try using subprocess to run nslookup (more reliable with different DNS configurations)
        result = subprocess.run(['nslookup', hostname], capture_output=True, text=True, timeout=5)
        output = result.stdout
        
        # Check if the lookup was successful
        if "NXDOMAIN" in output or "can't find" in output:
            return None, "Not found in DNS"
        
        # Extract IP address from nslookup output
        ip_match = re.search(r'Address:\s+(\d+\.\d+\.\d+\.\d+)(?!#)', output)
        if ip_match:
            return ip_match.group(1), "Found in DNS"
            
        # Fallback to socket.gethostbyname if nslookup doesn't provide clear results
        try:
            ip = socket.gethostbyname(hostname)
            return ip, "Found in DNS"
        except socket.gaierror:
            return None, "Not found in DNS"
            
    except subprocess.TimeoutExpired:
        return None, "DNS lookup timeout"
    except Exception as e:
        print(f"      DNS lookup error: {str(e)}")
        return None, "DNS lookup error"

def is_location_compliant(location, expected_location):
    """
    Check if the location string matches the expected location
    
    Args:
        location (str): Location string from API
        expected_location (str): Expected location string built from column references
        
    Returns:
        bool: True if compliant, False otherwise
    """
    if not location or not expected_location:
        return False
    
    # Remove any extra whitespace and convert to lowercase for comparison
    location = location.strip().lower()
    expected_location = expected_location.strip().lower()
    
    return location == expected_location

def process_excel_file(excel_path, api_url, api_token, location_format, domain_suffix=".sac.ragingwire.net", device_column=0):
    """
    Process the Excel file, query the API for each device, and update the Excel file
    
    Args:
        excel_path (str): Path to the Excel file
        api_url (str): Base URL for the LibreNMS API
        api_token (str): API token for authentication
        location_format (str): Format template with column references (e.g., $B.$C$E.$F)
        domain_suffix (str): Domain suffix to append to hostnames
        device_column (int): Zero-based index of the column containing device names
    """
    # Check if file exists
    if not os.path.exists(excel_path):
        print(f"Error: File {excel_path} not found")
        sys.exit(1)
    
    # Create a backup of the original file
    backup_path = f"{excel_path}.bak"
    if os.path.exists(excel_path):
        import shutil
        shutil.copy2(excel_path, backup_path)
        print(f"Backup created at {backup_path}")
    
    # Initialize LibreNMS client
    client = LibreNMSClient(api_url, api_token)
    
    # Read the Excel file - get all sheet names
    try:
        # Get all sheet names
        excel_file = pd.ExcelFile(excel_path)
        sheet_names = excel_file.sheet_names
        
        if not sheet_names:
            print(f"Error: Excel file {excel_path} has no sheets")
            sys.exit(1)
            
        print(f"Found {len(sheet_names)} sheets in the Excel file: {', '.join(sheet_names)}")
        
        # Create a writer to save to Excel with formatting
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Process each sheet
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            total_sheets = len(sheet_names)
            
            for sheet_idx, sheet_name in enumerate(sheet_names):
                print(f"\nProcessing sheet {sheet_idx+1}/{total_sheets}: '{sheet_name}'")
                
                # Read the sheet
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                
                # Check if the DataFrame is empty
                if df.empty:
                    print(f"  Warning: Sheet '{sheet_name}' is empty, skipping")
                    continue
                    
                # Check if there are any columns
                if len(df.columns) == 0:
                    print(f"  Warning: Sheet '{sheet_name}' has no columns, skipping")
                    continue
                
                # Identify the column containing device names
                if device_column < 0 or device_column >= len(df.columns):
                    print(f"  Warning: Device column index {device_column} is out of range. Using first column instead.")
                    device_col = df.columns[0]
                else:
                    device_col = df.columns[device_column]
                
                print(f"  Using column '{device_col}' for device names")
                
                # Define the new columns to add
                new_columns = ['hostname', 'ip', 'sysDescr', 'hardware', 'os', 'version', 
                              'last_polled', 'location', 'Expected_Location', 'Compliant', 'Status', 'DNS_IP', 'DNS_Status']
                
                # Add new columns to the DataFrame
                for col in new_columns:
                    if col not in df.columns:
                        df[col] = None
                
                # Process each row
                total_rows = len(df)
                processed_count = 0
                found_count = 0
                not_found_count = 0
                compliant_count = 0
                non_compliant_count = 0
                
                print(f"  Processing {total_rows} devices...")
                
                for idx, row in df.iterrows():
                    device_name = str(row[device_col]).strip()
                    if not device_name or pd.isna(device_name):
                        continue
                    
                    processed_count += 1
                    progress = (processed_count / total_rows) * 100
                    if processed_count % 5 == 0 or processed_count == total_rows:
                        print(f"  Progress: {processed_count}/{total_rows} ({progress:.1f}%)")
                    
                    # Handle domain suffix
                    # First, check if device_name already has any domain suffix
                    if '.' in device_name:
                        hostname_parts = device_name.split('.')
                        # If it has fewer than 2 parts, it's not a valid FQDN
                        if len(hostname_parts) < 2:
                            # Add the domain suffix
                            full_hostname = f"{device_name}{domain_suffix}"
                        else:
                            # Use the device name as is, assuming it already has a domain
                            full_hostname = device_name
                    else:
                        # No dots in the name, add the domain suffix
                        full_hostname = f"{device_name}{domain_suffix}"
                    
                    # Build expected location from column references
                    expected_location = ''
                    if location_format:
                        expected_location = build_expected_location(row, df.columns.tolist(), location_format)
                        df.at[idx, 'Expected_Location'] = expected_location if expected_location else ''
                    
                    # Query the API
                    print(f"    Querying API for device: {full_hostname}")
                    device_info = client.get_device_info(full_hostname)
                    
                    # Update the row with the API response data
                    if device_info:
                        found_count += 1
                        df.at[idx, 'hostname'] = device_info.get('hostname', '')
                        df.at[idx, 'ip'] = device_info.get('ip', '')
                        df.at[idx, 'sysDescr'] = device_info.get('sysDescr', '')
                        df.at[idx, 'hardware'] = device_info.get('hardware', '')
                        df.at[idx, 'os'] = device_info.get('os', '')
                        df.at[idx, 'version'] = device_info.get('version', '')
                        df.at[idx, 'last_polled'] = device_info.get('last_polled', '')
                        
                        # Get location and check compliance
                        location = device_info.get('location', '')
                        df.at[idx, 'location'] = location
                        
                        # Check if location is compliant with the expected format
                        is_compliant = False
                        if expected_location:
                            is_compliant = is_location_compliant(location, expected_location)
                            if is_compliant:
                                compliant_count += 1
                            else:
                                non_compliant_count += 1
                        
                        df.at[idx, 'Compliant'] = 'Yes' if is_compliant else 'No'
                        df.at[idx, 'Status'] = 'Found'
                        
                        print(f"      Found device information")
                        print(f"      Location: {location}")
                        if expected_location:
                            print(f"      Expected Location: {expected_location}")
                            print(f"      Compliant: {'Yes' if is_compliant else 'No'}")
                    else:
                        not_found_count += 1
                        df.at[idx, 'Status'] = 'Not found in LibreNMS'
                        print(f"      Device not found in LibreNMS")
                        
                        # Try DNS lookup for devices not found in LibreNMS
                        print(f"      Attempting DNS lookup for {full_hostname}...")
                        dns_ip, dns_status = perform_dns_lookup(full_hostname)
                        df.at[idx, 'DNS_IP'] = dns_ip
                        df.at[idx, 'DNS_Status'] = dns_status
                        print(f"      DNS lookup result: {dns_status} {dns_ip if dns_ip else ''}")
                
                # Save the DataFrame to the Excel file
                print(f"  Saving sheet '{sheet_name}' to Excel file...")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting to the sheet
                worksheet = writer.sheets[sheet_name]
                
                # Define styles
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                not_found_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                # Format headers
                for col_idx, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Format data rows
                for row_idx, row in enumerate(df.itertuples(), 2):  # Start from row 2 (after header)
                    # Get the status and compliant values
                    status_idx = df.columns.get_loc('Status') + 1 if 'Status' in df.columns else None
                    compliant_idx = df.columns.get_loc('Compliant') + 1 if 'Compliant' in df.columns else None
                    
                    if status_idx and hasattr(row, 'Status'):
                        status_cell = worksheet.cell(row=row_idx, column=status_idx)
                        status_value = getattr(row, 'Status', '')
                        
                        if status_value == 'Not found in LibreNMS':
                            # Color the entire row for not found devices
                            for col_idx in range(1, len(df.columns) + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = not_found_fill
                    
                    if compliant_idx and hasattr(row, 'Compliant'):
                        compliant_cell = worksheet.cell(row=row_idx, column=compliant_idx)
                        compliant_value = getattr(row, 'Compliant', '')
                        
                        if compliant_value == 'Yes':
                            compliant_cell.fill = compliant_fill
                        elif compliant_value == 'No':
                            compliant_cell.fill = non_compliant_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Print summary for this sheet
                print(f"\n  Summary for sheet '{sheet_name}':\n")
                print(f"    Total devices processed: {processed_count}")
                print(f"    Devices found in LibreNMS: {found_count}")
                print(f"    Devices not found in LibreNMS: {not_found_count}")
                if location_format:
                    print(f"    Devices with compliant location: {compliant_count}")
                    print(f"    Devices with non-compliant location: {non_compliant_count}")
            
            # Create a summary sheet
            print(f"\nCreating summary sheet...")
            
            # Create a summary DataFrame
            summary_data = {
                'Sheet Name': [],
                'Total Devices': [],
                'Devices Found': [],
                'Devices Not Found': [],
                'Compliant Locations': [],
                'Non-Compliant Locations': [],
                'Processed Date': []
            }
            
            # Collect summary data from all sheets
            for sheet_name in sheet_names:
                # Read the sheet again to get the latest data
                try:
                    sheet_df = pd.read_excel(excel_path, sheet_name=sheet_name)
                    
                    # Skip empty sheets
                    if sheet_df.empty or len(sheet_df.columns) == 0:
                        continue
                    
                    # Count devices
                    total_devices = len(sheet_df)
                    
                    # Count found/not found devices
                    if 'Status' in sheet_df.columns:
                        found_devices = len(sheet_df[sheet_df['Status'] == 'Found'])
                        not_found_devices = len(sheet_df[sheet_df['Status'] == 'Not found in LibreNMS'])
                    else:
                        found_devices = 0
                        not_found_devices = 0
                    
                    # Count compliant/non-compliant locations
                    if 'Compliant' in sheet_df.columns:
                        compliant_locations = len(sheet_df[sheet_df['Compliant'] == 'Yes'])
                        non_compliant_locations = len(sheet_df[sheet_df['Compliant'] == 'No'])
                    else:
                        compliant_locations = 0
                        non_compliant_locations = 0
                    
                    # Add to summary data
                    summary_data['Sheet Name'].append(sheet_name)
                    summary_data['Total Devices'].append(total_devices)
                    summary_data['Devices Found'].append(found_devices)
                    summary_data['Devices Not Found'].append(not_found_devices)
                    summary_data['Compliant Locations'].append(compliant_locations)
                    summary_data['Non-Compliant Locations'].append(non_compliant_locations)
                    summary_data['Processed Date'].append(pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'))
                except Exception as e:
                    print(f"  Warning: Could not read sheet '{sheet_name}' for summary: {str(e)}")
            
            # Create summary DataFrame
            summary_df = pd.DataFrame(summary_data)
            
            # Add a row with totals
            totals = {
                'Sheet Name': 'TOTAL',
                'Total Devices': summary_df['Total Devices'].sum(),
                'Devices Found': summary_df['Devices Found'].sum(),
                'Devices Not Found': summary_df['Devices Not Found'].sum(),
                'Compliant Locations': summary_df['Compliant Locations'].sum(),
                'Non-Compliant Locations': summary_df['Non-Compliant Locations'].sum(),
                'Processed Date': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            summary_df = pd.concat([summary_df, pd.DataFrame([totals])], ignore_index=True)
            
            # Save summary to a new sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the summary sheet
            summary_worksheet = writer.sheets['Summary']
            
            # Define styles
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            total_font = Font(bold=True)
            
            # Format headers
            for col_idx, column in enumerate(summary_df.columns, 1):
                cell = summary_worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
            
            # Format totals row
            for col_idx in range(1, len(summary_df.columns) + 1):
                cell = summary_worksheet.cell(row=len(summary_df) + 1, column=col_idx)
                cell.fill = total_fill
                cell.font = total_font
            
            # Auto-adjust column widths
            for column in summary_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                summary_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"Summary sheet created")
            print(f"\nAll sheets processed and saved to {excel_path}")
            
    except PermissionError:
        print(f"Error: Permission denied when accessing {excel_path}. Make sure the file is not open in another program.")
        sys.exit(1)
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

def main():
    """Main function to parse arguments and run the script"""
    parser = argparse.ArgumentParser(description='SNMP Location Lookup Tool')
    parser.add_argument('--excel', required=True, help='Path to the Excel file')
    parser.add_argument('--api-url', required=True, help='LibreNMS API URL (e.g., https://10.1.0.183)')
    parser.add_argument('--api-token', required=True, help='LibreNMS API token')
    parser.add_argument('--location-format', default='$B.$C.$D.$E', 
                        help='Format template with column references (e.g., $B.$C$E.$F) (default: $B.$C.$D.$E)')
    parser.add_argument('--domain-suffix', default='.sac.ragingwire.net', help='Domain suffix to append to hostnames')
    parser.add_argument('--device-column', type=int, default=0, 
                        help='Zero-based index of the column containing device names (default: 0, which is column A)')
    
    args = parser.parse_args()
    
    # Validate arguments
    if not args.excel.endswith(('.xlsx', '.xls')):
        print("Warning: The specified file does not have an Excel extension (.xlsx or .xls)")
        response = input("Continue anyway? (y/n): ")
        if response.lower() != 'y':
            sys.exit(0)
    
    process_excel_file(args.excel, args.api_url, args.api_token, args.location_format, args.domain_suffix, args.device_column)

if __name__ == "__main__":
    main()
